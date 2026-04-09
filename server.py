# File: server.py
from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
import cv2
import numpy as np
import base64
import os
import time
import traceback
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Import các module từ dự án
from src.face_model import init_insightface, extract_faces, extract_embedding
from src.database import load_database, find_best_match, add_new_face_to_db
from src.liveness_check import check_liveness, MODEL_AVAILABLE
from src.checkin import log_checkin, get_work_status, init_checkin_log
from src.config import Config

app = Flask(__name__)
CORS(app)

print("="*60)
print("ĐANG KHỞI TẠO SERVER AI...")
print("="*60)

try:
    face_app = init_insightface()
    db_embeddings, db_labels_df = load_database()
    init_checkin_log()
    print(f"[SUCCESS] Đã nạp Database: {len(db_embeddings)} người.")
    if not MODEL_AVAILABLE:
        print("[WARN] Module Liveness chưa được kích hoạt!")
    else:
        print("[SUCCESS] Module Liveness đã sẵn sàng.")
except Exception as e:
    print(f"[ERROR] Lỗi khởi tạo hệ thống: {e}")
    traceback.print_exc()
    exit(1)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/checkin', methods=['POST'])
def api_checkin():
    start_time = time.time()
    data = request.json
    
    if not data:
        return jsonify({'success': False, 'message': 'Dữ liệu rỗng'}), 400

    image_data = data.get('image')
    mode = data.get('mode', 'checkin')
    reg_id = data.get('reg_id')
    reg_name = data.get('reg_name')

    print(f"\n{'='*50}")
    print(f"[DEBUG] Nhận yêu cầu MODE: {mode.upper()}")
    print(f"{'='*50}")

    try:
        # --- 1. GIẢI MÃ ẢNH ---
        if not image_data or not isinstance(image_data, str):
            return jsonify({'success': False, 'message': 'Không nhận được ảnh hợp lệ'}), 400

        try:
            if ',' in image_data:
                header, encoded = image_data.split(',', 1)
            else:
                encoded = image_data
            
            img_bytes = base64.b64decode(encoded)
            nparr = np.frombuffer(img_bytes, np.uint8)
            frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            
            if frame is None or frame.size == 0:
                return jsonify({'success': False, 'message': 'Không thể giải mã ảnh'}), 400
        except Exception as e:
            return jsonify({'success': False, 'message': f'Lỗi xử lý ảnh: {str(e)}'}), 400

        # --- 2. PHÁT HIỆN KHUÔN MẶT ---
        faces = extract_faces(face_app, frame)
        if len(faces) == 0:
            return jsonify({'success': False, 'message': 'Không tìm thấy khuôn mặt.'}), 404

        face = max(faces, key=lambda x: (x.bbox[2]-x.bbox[0]) * (x.bbox[3]-x.bbox[1]))
        x1, y1, x2, y2 = map(int, face.bbox)
        face_crop = frame[y1:y2, x1:x2]
        
        if face_crop.size == 0:
            return jsonify({'success': False, 'message': 'Khuôn mặt quá nhỏ'}), 400

        # --- 3. XỬ LÝ LOGIC ---
        if mode == 'register':
            if not reg_id or not reg_name:
                return jsonify({'success': False, 'message': 'Thiếu thông tin Mã hoặc Tên'}), 400
            
            print(f"[REGISTER] Đang đăng ký mới: {reg_name} ({reg_id})...")
            embedding = extract_embedding(face)
            
            try:
                global db_embeddings, db_labels_df
                add_new_face_to_db(reg_id, reg_name, embedding)
                db_embeddings, db_labels_df = load_database()
                
                current_time = datetime.now().strftime("%H:%M:%S - %d/%m/%Y")
                return jsonify({
                    'success': True, 'message': f'Đã đăng ký thành công cho {reg_name}',
                    'id': reg_id, 'name': reg_name, 'timestamp': current_time
                })
            except Exception as e:
                return jsonify({'success': False, 'message': f'Lỗi lưu dữ liệu: {str(e)}'}), 500

        else:
            is_real, liveness_score = check_liveness(face_crop)
            if not is_real:
                return jsonify({
                    'success': False, 'message': 'PHÁT HIỆN GIẢ MẠO!',
                    'type': 'SPOOF', 'score': f"{liveness_score:.2f}"
                }), 403

            embedding = extract_embedding(face)
            match_label, similarity = find_best_match(embedding, db_embeddings, db_labels_df)

            if match_label:
                pid = match_label["ID_Name"]
                name = match_label["Name"]
                
                if mode == 'checkin': log_type = "CHECK_IN"
                elif mode == 'checkout': log_type = "CHECK_OUT"
                else: log_type = get_work_status(datetime.now())
                
                current_time = datetime.now()
                time_str = current_time.strftime("%H:%M:%S - %d/%m/%Y")
                
                log_checkin(pid, name, similarity, log_type)
                
                return jsonify({
                    'success': True, 'message': f'{log_type.replace("_", "-")} thành công!',
                    'name': name, 'id': pid, 'type': log_type,
                    'similarity': f"{similarity:.2f}", 'timestamp': time_str
                })
            else:
                return jsonify({
                    'success': False, 'message': 'Khuôn mặt không tồn tại.', 'type': 'UNKNOWN'
                }), 404

    except Exception as e:
        print(f"[ERROR] Lỗi hệ thống: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Lỗi server: {str(e)}'}), 500

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    """Xuất file Excel báo cáo chấm công"""
    try:
        if not os.path.exists(Config.checkin_log):
            return jsonify({'success': False, 'message': 'Chưa có dữ liệu.'}), 404
        
        df = pd.read_csv(Config.checkin_log)
        if df.empty:
            return jsonify({'success': False, 'message': 'Dữ liệu trống.'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Bang_Cham_Cong"

        headers = ["STT", "Thời Gian", "Mã NV/SV", "Họ Tên", "Loại", "Độ Tin Cậy", "Ghi Chú"]
        ws.append(headers)

        # Format Header
        header_fill = PatternFill(start_color="4e54c8", end_color="4e54c8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        # Điền dữ liệu
        for idx, row in df.iterrows():
            time_val = row.get('Time', '')
            id_val = row.get('ID_Name', '')
            name_val = row.get('Name', '')
            type_val = row.get('Type', '')
            sim_val = row.get('Similarity', '')
            
            # --- SỬA LỖI Ở ĐÂY: Chuyển đổi sim_val sang float an toàn ---
            try:
                sim_float = float(sim_val) if sim_val is not None else 0.0
                sim_formatted = f"{sim_float:.2f}"
            except (ValueError, TypeError):
                sim_formatted = str(sim_val) # Fallback nếu không chuyển được
            # -----------------------------------------------------------
            
            fill_color = None
            if type_val == 'CHECK_IN': fill_color = "d1fae5"
            elif type_val == 'CHECK_OUT': fill_color = "fee2e2"
            
            new_row = [idx + 1, time_val, id_val, name_val, type_val, sim_formatted, "Bình thường"]
            ws.append(new_row)

            row_idx = idx + 2
            for cell in ws[row_idx]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                if fill_color and cell.column_letter != 'A':
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        column_widths = [5, 20, 15, 30, 15, 12, 20]
        for i, width in enumerate(column_widths):
            col_letter = get_column_letter(i + 1)
            ws.column_dimensions[col_letter].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Bao_Cao_Cham_Cong_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"[ERROR] Lỗi xuất Excel: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Lỗi server: {str(e)}'}), 500

if __name__ == '__main__':
    print("\nSERVER ĐANG CHẠY")
    app.run(host='0.0.0.0', port=5000, debug=False)